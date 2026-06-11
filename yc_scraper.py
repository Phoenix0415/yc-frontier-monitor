#!/usr/bin/env python3
"""
YC batch scraper -> Excel.

Run at home (unrestricted network). Examples:
    python yc_scraper.py                       # defaults to "Fall 2025"
    python yc_scraper.py "Winter 2026"         # any single batch
    python yc_scraper.py "Spring 2026" "Summer 2026"   # several -> one file

Install once:  pip install requests openpyxl

Output: yc_<batch>.xlsx in the current folder.

Two-step pull:
  1) YC's directory is backed by an Algolia search index -> one fast call gets
     name, bio, industry, batch, founded, location, status, team size, hiring,
     website, and the long description (where revenue usually hides).
  2) Founder count isn't in that index, so we open each company page and read it
     from the embedded page data. This is the slower part but still a couple min.
"""

import sys
import re
import json
import time
import html as htmllib
from datetime import datetime, timezone
from urllib.parse import urlencode
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

UA = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0 Safari/537.36"}

ALGOLIA_INDEX = "YCCompany_production"
APP_ID_FALLBACK = "45BWZJ1SGC"   # public; only used if auto-discovery fails

# ----- column layout (this order is the agreed one; groups noted) -----
# header, width, group
COLUMNS = [
    ("Name",            22, "id"),
    ("One-line bio",    50, "id"),
    ("Industry",        30, "id"),
    ("Batch",           13, "id"),
    ("Founded",         10, "id"),
    ("Location",        18, "id"),
    ("Status",          11, "id"),
    ("Team size",       11, "team"),     # team-size and founder-count sit
    ("Founder count",   13, "team"),     #   side by side, on purpose
    ("Revenue / ARR",   30, "traction"),
    ("Hiring",          11, "activity"),
    ("Website",         34, "link"),
    ("Full intro",      70, "detail"),
]
GROUP_FILL = {
    "id":       "2F6BFF",   # accent blue
    "team":     "1E40AF",   # darker blue so the pair stands out at a glance
    "traction": "2F6BFF",
    "activity": "2F6BFF",
    "link":     "2F6BFF",
    "detail":   "2F6BFF",
}

# ---------------------------------------------------------------- Algolia
def get_algolia_creds():
    """Find the public Algolia app id + search-only key from the YC site."""
    txt = requests.get("https://www.ycombinator.com/companies",
                       headers=UA, timeout=30).text
    # Current site (Vite/Inertia rebuild) inlines the creds as a JSON blob:
    #   window.AlgoliaOpts = {"app":"45BWZJ1SGC","key":"<base64 search key>"}
    # The key is a secured base64 key, not 32-hex, so the old regex misses it.
    m = re.search(r'window\.AlgoliaOpts\s*=\s*(\{.*?\})\s*;', txt, re.S)
    if m:
        try:
            opts = json.loads(m.group(1))
            if opts.get("app") and opts.get("key"):
                return opts["app"], opts["key"]
        except Exception:
            pass
    # Legacy Next.js layout: inline algoliaApiKey/algoliaAppId.
    key = re.search(r'algoliaApiKey:"([0-9a-f]{32,})"', txt)
    app = re.search(r'algoliaAppId:"([A-Z0-9]{8,})"', txt)
    if key and app:
        return app.group(1), key.group(1)
    # scan the JS chunks the page loads
    for path in set(re.findall(r'/_next/static/[^"\']+\.js', txt)):
        try:
            js = requests.get("https://www.ycombinator.com" + path,
                              headers=UA, timeout=30).text
        except Exception:
            continue
        key = re.search(r'algoliaApiKey:"([0-9a-f]{32,})"', js)
        app = re.search(r'algoliaAppId:"([A-Z0-9]{8,})"', js)
        if key and app:
            return app.group(1), key.group(1)
    raise SystemExit(
        "Could not auto-find the Algolia key.\n"
        "Open https://www.ycombinator.com/companies in Chrome, DevTools > Network,\n"
        "filter 'algolia', click any request, and copy the two request headers:\n"
        "  X-Algolia-Application-Id  and  X-Algolia-API-Key\n"
        "Then hard-code them in main() (app_id, api_key) and re-run.")

def fetch_batch(app_id, api_key, batch):
    url = f"https://{app_id.lower()}-dsn.algolia.net/1/indexes/{ALGOLIA_INDEX}/query"
    headers = {"X-Algolia-API-Key": api_key,
               "X-Algolia-Application-Id": app_id,
               "Content-Type": "application/json"}
    hits, page, pages = [], 0, 1
    while page < pages:
        params = urlencode({
            "hitsPerPage": 1000,
            "page": page,
            "facetFilters": json.dumps([[f"batch:{batch}"]]),
        })
        r = requests.post(url, headers=headers, json={"params": params}, timeout=30)
        r.raise_for_status()
        data = r.json()
        hits += data.get("hits", [])
        pages = data.get("nbPages", 1)
        page += 1
    return hits

# ---------------------------------------------------------------- per-company
def deep_find(obj, key):
    out = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            if k == key:
                out.append(v)
            out += deep_find(v, key)
    elif isinstance(obj, list):
        for it in obj:
            out += deep_find(it, key)
    return out

def founder_count(slug):
    """Open the company page and count founders from embedded page data."""
    try:
        html = requests.get(f"https://www.ycombinator.com/companies/{slug}",
                            headers=UA, timeout=30).text
    except Exception:
        return ""
    # The page is now an Inertia app: the company payload is HTML-escaped JSON
    # in the root element's data-page attribute (the old __NEXT_DATA__ script
    # is gone). Founders live at props.company.founders.
    m = re.search(r'data-page="(.*?)"\s*>', html, re.S)
    if m:
        try:
            data = json.loads(htmllib.unescape(m.group(1)))
            lists = [v for v in deep_find(data, "founders")
                     if isinstance(v, list) and v and isinstance(v[0], dict)]
            if lists:
                return len(max(lists, key=len))
        except Exception:
            pass
    return ""   # leave blank rather than guess wrong

# ---------------------------------------------------------------- revenue
# A sentence is "any run of chars up to a sentence-ending period". A period that
# is part of a number (e.g. "6.5x", "$1,250,000.50") is followed by a digit, so
# `\.(?=\d)` keeps it inside the sentence; only a period NOT followed by a digit
# ends the match. Old version used [^.]* and truncated on the first decimal point.
NOTDOT = r'(?:[^.]|\.(?=\d))'
REV = re.compile(
    NOTDOT + r'*\$\s?[\d.,]+\s?(?:[KMB]|thousand|million|billion)?\+?' + NOTDOT + r'*?'
    r'(?:ARR|MRR|revenue|run[\s-]?rate|GMV|bookings)' + NOTDOT + r'*\.',
    re.I)

def extract_revenue(text):
    if not text:
        return ""
    m = REV.search(text)
    if not m:
        return ""
    snippet = re.sub(r'\s+', ' ', m.group(0)).strip()
    i = snippet.find("$")           # drop any lead-in clause before the figure
    return snippet[i:].strip(" .,") if i > 0 else snippet

# ---------------------------------------------------------------- mapping
def first(*vals, default=""):
    for v in vals:
        if v not in (None, "", []):
            return v
    return default

def as_year(v):
    """Normalize a founding-year value to a 4-digit year string.

    YC dropped the old `year_founded` field; only `launched_at` (a unix epoch)
    now survives in the index, so the raw fallback would dump a timestamp.
    Convert an epoch to its year; pass through values that already look like a
    year. Keeps the 'Founded' column meaning a year, as it always has."""
    if v in (None, "", []):
        return ""
    try:
        n = int(v)
    except (TypeError, ValueError):
        return str(v)
    if n > 9999:   # unix epoch seconds -> calendar year
        return str(datetime.fromtimestamp(n, tz=timezone.utc).year)
    return str(n)  # already a plain year

def map_hit(hit):
    tags = first(hit.get("tags"), hit.get("industries"), default=[])
    if isinstance(tags, str):
        tags = [tags]
    loc = first(hit.get("all_locations"), hit.get("location"),
                hit.get("locations"), default="")
    if isinstance(loc, list):
        loc = loc[0] if loc else ""
    hiring = hit.get("isHiring")
    desc = first(hit.get("long_description"), hit.get("description"))
    return {
        "Name":          first(hit.get("name")),
        "One-line bio":  first(hit.get("one_liner"), hit.get("oneLiner")),
        "Industry":      "; ".join(tags) if isinstance(tags, list) else str(tags),
        "Batch":         first(hit.get("batch")),
        "Founded":       as_year(first(hit.get("year_founded"), hit.get("launched_at"))),
        "Location":      str(loc).split(",")[0].strip() if loc else "",
        "Status":        first(hit.get("status")),
        "Team size":     first(hit.get("team_size"), default=""),
        "Revenue / ARR": extract_revenue(desc),
        "Hiring":        "Yes" if hiring else ("No" if hiring is not None else ""),
        "Website":       first(hit.get("website")),
        "Full intro":    re.sub(r'\s+\n', '\n', desc).strip() if desc else "",
        "_slug":         first(hit.get("slug")),
    }

# ---------------------------------------------------------------- excel
def write_excel(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "YC"
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, (header, width, group) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=GROUP_FILL[group])
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = border
        ws.column_dimensions[chr(64 + c)].width = width

    for r, row in enumerate(rows, start=2):
        for c, (header, _, _) in enumerate(COLUMNS, start=1):
            value = row.get(header, "")
            cell = ws.cell(row=r, column=c, value=value)
            if header == "Website" and value:
                # one-click clickable link; show the URL as the label
                cell.hyperlink = value
                cell.font = Font(name="Arial", size=10, color="2F6BFF",
                                 underline="single")
            else:
                cell.font = Font(name="Arial", size=10)
            wrap = header in ("One-line bio", "Revenue / ARR", "Industry",
                              "Full intro")
            cell.alignment = Alignment(vertical="top", wrap_text=wrap,
                                       horizontal="left")
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(COLUMNS))}{len(rows) + 1}"
    ws.row_dimensions[1].height = 30
    wb.save(path)

# ---------------------------------------------------------------- main
def main():
    batches = sys.argv[1:] or ["Fall 2025"]
    app_id, api_key = get_algolia_creds()
    print(f"Algolia OK ({app_id}). Pulling: {', '.join(batches)}")

    hits = []
    for b in batches:
        h = fetch_batch(app_id, api_key, b)
        print(f"  {b}: {len(h)} companies")
        hits += h

    rows = [map_hit(h) for h in hits]

    # founder count, concurrently
    print("Reading founder counts...")
    def job(row):
        row["Founder count"] = founder_count(row["_slug"]) if row["_slug"] else ""
        return row
    with ThreadPoolExecutor(max_workers=8) as ex:
        futs = [ex.submit(job, row) for row in rows]
        done = 0
        for _ in as_completed(futs):
            done += 1
            if done % 25 == 0:
                print(f"  {done}/{len(rows)}")

    rows.sort(key=lambda r: (r.get("Name") or "").lower())
    for r in rows:
        r.pop("_slug", None)

    name = "yc_" + "_".join(b.replace(" ", "") for b in batches) + ".xlsx"
    write_excel(rows, name)
    print(f"Done -> {name}  ({len(rows)} rows)")

if __name__ == "__main__":
    main()
